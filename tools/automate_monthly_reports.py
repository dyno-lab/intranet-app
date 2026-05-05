from __future__ import annotations

import argparse
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(r"F:\FARO\Automatizaciones")
MONTH_NAMES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}

# Nombre de archivo recibido -> nombre de fila en los resumenes anuales.
RESIDENTIAL_BY_FILE = {
    "achavier": "Arístides Chavier",
    "pjr": "Pedro J. Rosaly",
    "jpl": "Juan Ponce de León",
    "era": "Ernesto Ramos Antonini",
    "rln": "Rafael López Nussa",
    "ceiba": "La Ceiba",
    "ls": "Leónardo Santiago",
    "vg": "Valles de Guayama",
    "jg": "Jardines de Guamani",
    "fc": "Fernando Calimano",
    "carioca": "San Antonio Carioca",
    "carmen": "El Carmen",
    "mhr": "Manuel Hernandez Rosa",
    "rh": "Rafael Hernandez",
    "cl": "Columbus Landing",
}


def norm(value: str | None) -> str:
    return "" if value is None else " ".join(str(value).strip().lower().split())


def copy_sheet_template(wb, source_name: str, target_name: str):
    source = wb[source_name]
    target = wb.copy_worksheet(source)
    target.title = target_name
    return target


def ensure_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name], False
    return copy_sheet_template(wb, wb.sheetnames[0], sheet_name), True


def find_row_by_residential(ws, residential: str) -> int | None:
    wanted = norm(residential)
    for row in range(10, 28):
        if norm(ws.cell(row, 2).value) == wanted:
            return row
    return None


def backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = path.with_name(f"{path.stem}.bak_{stamp}{path.suffix}")
    shutil.copy2(path, out)
    return out


def read_source_metrics(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    emb = wb["Embarazo"]
    des = wb["DesercionEscolar"]

    embarazo = {
        "female": int(emb["E88"].value or 0),
        "male": int(emb["F88"].value or 0),
        "pregnant_female": 0,
        "impregnated_male": 0,
    }
    for row in range(8, 87):
        is_female = str(emb.cell(row, 4).value or "").strip().upper() == "X"
        is_male = str(emb.cell(row, 5).value or "").strip().upper() == "X"
        pregnancy_yes = str(emb.cell(row, 10).value or "").strip().upper() == "X"
        if pregnancy_yes and is_female:
            embarazo["pregnant_female"] += 1
        if pregnancy_yes and is_male:
            embarazo["impregnated_male"] += 1

    desercion = {
        "female": int(des["R8"].value or 0),
        "male": int(des["S8"].value or 0),
        "grades": [int(des.cell(8, col).value or 0) for col in range(20, 35)],  # SC, EE, K, 1..12
        "tutorias": int(des["AI8"].value or 0),
        "escuela": int(des["AJ8"].value or 0),
        "notas": [int(des.cell(8, col).value or 0) for col in range(37, 41)],  # 10,20,30,40 sem
    }
    wb.close()
    return {"embarazo": embarazo, "desercion": desercion}


def set_title(ws, title: str):
    ws["B7"] = title


def fill_embarazo(ws, row: int, data: dict):
    ws.cell(row, 4).value = data["female"]
    ws.cell(row, 5).value = data["male"]
    ws.cell(row, 6).value = data["pregnant_female"]
    ws.cell(row, 7).value = data["impregnated_male"]
    # C y H conservan formulas del template.


def fill_desercion(ws, row: int, data: dict):
    ws.cell(row, 4).value = data["female"]
    ws.cell(row, 5).value = data["male"]
    for idx, value in enumerate(data["grades"], start=6):
        ws.cell(row, idx).value = value
    ws.cell(row, 21).value = data["tutorias"]
    ws.cell(row, 23).value = data["escuela"]
    for idx, value in enumerate(data["notas"], start=25):
        ws.cell(row, idx).value = value
    # C, V y X conservan formulas del template.


def main() -> int:
    parser = argparse.ArgumentParser(description="Llena Embarazo y eDesercion desde informes mensuales XLSM.")
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, required=True)
    parser.add_argument("--write", action="store_true", help="Escribe cambios reales. Sin esto solo muestra dry-run.")
    args = parser.parse_args()

    month_name = MONTH_NAMES[args.month]
    sheet_name = f"{month_name}{args.year}"
    source_dir = BASE / "Informe Mensuales" / str(args.year) / f"{month_name} {args.year}" / "ya"
    embarazo_path = BASE / "Embarazo" / str(args.year) / f"Embarazo{args.year}.xlsx"
    desercion_path = BASE / "eDesercion" / str(args.year) / f"Desercion{args.year}.xlsx"

    if not source_dir.exists():
        raise SystemExit(f"No existe carpeta fuente: {source_dir}")

    source_files = sorted(source_dir.glob("*.xlsm"))
    print(f"Fuente: {source_dir}")
    print(f"Archivos encontrados: {len(source_files)}")
    print(f"Hoja destino: {sheet_name}")

    wb_emb = load_workbook(embarazo_path)
    wb_des = load_workbook(desercion_path)
    ws_emb, emb_created = ensure_sheet(wb_emb, sheet_name)
    ws_des, des_created = ensure_sheet(wb_des, sheet_name)
    set_title(ws_emb, f"Prevención de Embarazos en Jóvenes Féminas y Masculinos ({month_name} {args.year})")
    set_title(ws_des, f"Deserción Escolar ({month_name} {args.year})")

    processed = []
    warnings = []
    for source_file in source_files:
        key = source_file.stem.lower()
        residential = RESIDENTIAL_BY_FILE.get(key)
        if not residential:
            warnings.append(f"Sin mapeo para archivo {source_file.name}")
            continue
        row_emb = find_row_by_residential(ws_emb, residential)
        row_des = find_row_by_residential(ws_des, residential)
        if not row_emb or not row_des:
            warnings.append(f"No encontre fila destino para {residential} ({source_file.name})")
            continue
        metrics = read_source_metrics(source_file)
        fill_embarazo(ws_emb, row_emb, metrics["embarazo"])
        fill_desercion(ws_des, row_des, metrics["desercion"])
        processed.append((source_file.name, residential, row_emb, row_des, metrics))

    target_residentials = [ws_emb.cell(row, 2).value for row in range(10, 28)]
    received_residentials = {item[1] for item in processed}
    missing = [name for name in target_residentials if name and name != "TOTAL " and name not in received_residentials]

    print("\nProcesados:")
    for filename, residential, row_emb, row_des, metrics in processed:
        emb = metrics["embarazo"]
        des = metrics["desercion"]
        print(f"- {filename}: {residential} | Embarazo F/M={emb['female']}/{emb['male']} preg={emb['pregnant_female']}/{emb['impregnated_male']} | Desercion F/M={des['female']}/{des['male']} tutorias={des['tutorias']} escuela={des['escuela']}")

    if missing:
        print("\nResidenciales sin archivo recibido / sin procesar:")
        for name in missing:
            print(f"- {name}")
    if warnings:
        print("\nAdvertencias:")
        for warning in warnings:
            print(f"- {warning}")

    if args.write:
        b1 = backup(embarazo_path)
        b2 = backup(desercion_path)
        wb_emb.save(embarazo_path)
        wb_des.save(desercion_path)
        print(f"\nESCRITO. Backups creados:\n- {b1}\n- {b2}")
    else:
        print("\nDRY-RUN: no se escribieron cambios. Usa --write para actualizar los Excel finales.")
        print(f"Hojas {'creadas' if emb_created or des_created else 'existentes'} en memoria: {sheet_name}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
