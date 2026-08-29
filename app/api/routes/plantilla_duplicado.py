from __future__ import annotations

from datetime import date
from io import BytesIO
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, status
from fastapi.responses import RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_db
from app.core.auth import require_admin
from app.core.residential_scope import has_global_residential_access, require_record_residential_id
from app.models.proposal import Proposal
from app.models.residential import Residential
from app.models.user import User
from app.services.consolidado_mensual_service import MONTH_NAMES
from app.services.plantilla_duplicado_service import build_plantilla_duplicado_context
from app.services.report_pdf import PDFBackendUnavailableError, PDFRenderError, render_template_to_pdf_bytes

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
MONTH_OPTIONS = [(idx, name.capitalize()) for idx, name in MONTH_NAMES.items()]
PERIOD_TYPE_OPTIONS = [
    {"value": "monthly", "label": "Mensual"},
    {"value": "custom", "label": "Personalizado"},
]


def _validate_month_year(month: int, year: int) -> None:
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Mes inválido.")
    if year < 2000 or year > 2100:
        raise HTTPException(status_code=400, detail="Año inválido.")


def _validate_period(period_type: str, month: int | None, year: int | None, start_date: str | None, end_date: str | None) -> None:
    if period_type == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="Debe indicar fecha inicial y fecha final.")
        try:
            parsed_start = date.fromisoformat(start_date)
            parsed_end = date.fromisoformat(end_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha inválido.")
        if parsed_start > parsed_end:
            raise HTTPException(status_code=400, detail="La fecha inicial no puede ser mayor que la fecha final.")
        return
    if month is None or year is None:
        raise HTTPException(status_code=400, detail="Debe indicar mes y año.")
    _validate_month_year(month, year)


def _parse_optional_int(value: str | int | None) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Filtro inválido.")


def _resolve_report_residential_id(
    request: Request,
    current_user: User,
    requested_residential_id: int | None,
) -> int | None:
    if has_global_residential_access(current_user):
        return requested_residential_id
    active_residential_id = require_record_residential_id(request, current_user)
    if requested_residential_id is not None and requested_residential_id != active_residential_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="El reporte solo puede consultarse para el residencial activo.",
        )
    return active_residential_id


def _base_context(db: Session, residential_id: int | None = None) -> dict:
    proposals = db.execute(select(Proposal).order_by(Proposal.code)).scalars().all()
    residential_stmt = (
        select(Residential)
        .where(Residential.is_active == True)  # noqa: E712
        .order_by(Residential.municipality, Residential.name)
    )
    if residential_id is not None:
        residential_stmt = residential_stmt.where(Residential.residential_id == residential_id)
    residentials = db.execute(residential_stmt).scalars().all()
    return {
        "proposals": proposals,
        "residentials": residentials,
        "month_options": MONTH_OPTIONS,
        "period_type_options": PERIOD_TYPE_OPTIONS,
        "year_options": list(range(date.today().year - 2, date.today().year + 3)),
    }


def _build_context(
    db: Session,
    current_user: User,
    month: int | None,
    year: int | None,
    proposal_id: int | None,
    residential_id: int | None,
    period_type: str = "monthly",
    start_date: str | None = None,
    end_date: str | None = None,
) -> dict:
    _validate_period(period_type, month, year, start_date, end_date)
    report = build_plantilla_duplicado_context(
        db,
        month=month,
        year=year,
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
        proposal_id=proposal_id,
        residential_id=residential_id,
        current_user=current_user,
    )
    residential_scope_id = None if has_global_residential_access(current_user) else residential_id
    return {**_base_context(db, residential_scope_id), **report, "current_user": current_user, "msg": None}


def _query_params(month: int | None, year: int | None, proposal_id: int | None, residential_id: int | None, period_type: str, start_date: str | None, end_date: str | None) -> str:
    params = {"period_type": period_type}
    if month:
        params["month"] = month
    if year:
        params["year"] = year
    if start_date:
        params["start_date"] = start_date
    if end_date:
        params["end_date"] = end_date
    if proposal_id:
        params["proposal_id"] = proposal_id
    if residential_id:
        params["residential_id"] = residential_id
    return urlencode(params)


def _download_filename(prefix: str, context: dict, extension: str) -> str:
    proposal_code = getattr(context.get("proposal"), "code", None)
    proposal_part = f"_{proposal_code}" if proposal_code else ""
    if context.get("selected_period_type") == "custom":
        suffix = f"{context.get('selected_start_date')}_a_{context.get('selected_end_date')}"
    else:
        suffix = f"{context['month']:02d}_{context['year']}"
    return f"{prefix}{proposal_part}_{suffix}.{extension}"


@router.get("/plantilla-duplicado")
def plantilla_duplicado_index(
    request: Request,
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    residential_id: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    selected_month = month or date.today().month
    selected_year = year or date.today().year
    selected_residential_id = _resolve_report_residential_id(
        request,
        current_user,
        _parse_optional_int(residential_id),
    )
    context = _build_context(
        db,
        current_user,
        selected_month,
        selected_year,
        _parse_optional_int(proposal_id),
        selected_residential_id,
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
    )
    context["request"] = request
    return templates.TemplateResponse("ui/admin/plantilla_duplicado.html", context)


@router.post("/plantilla-duplicado/generar")
def plantilla_duplicado_generar(
    request: Request,
    month: int | None = Form(None),
    year: int | None = Form(None),
    period_type: str = Form("monthly"),
    start_date: str | None = Form(None),
    end_date: str | None = Form(None),
    proposal_id: str | None = Form(None),
    residential_id: str | None = Form(None),
    output: str = Form("pdf"),
    current_user: User = Depends(require_admin),
):
    _validate_period(period_type, month, year, start_date, end_date)
    selected_residential_id = _resolve_report_residential_id(
        request,
        current_user,
        _parse_optional_int(residential_id),
    )
    target = "pdf" if output == "pdf" else "excel" if output == "excel" else ""
    if not target:
        raise HTTPException(status_code=400, detail="Salida inválida.")
    return RedirectResponse(
        f"/ui/admin/plantilla-duplicado/{target}?{_query_params(month, year, _parse_optional_int(proposal_id), selected_residential_id, period_type, start_date, end_date)}",
        status_code=status.HTTP_303_SEE_OTHER,
    )


@router.get("/plantilla-duplicado/pdf")
def plantilla_duplicado_pdf(
    request: Request,
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    residential_id: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    selected_residential_id = _resolve_report_residential_id(
        request,
        current_user,
        _parse_optional_int(residential_id),
    )
    context = _build_context(db, current_user, month, year, _parse_optional_int(proposal_id), selected_residential_id, period_type=period_type, start_date=start_date, end_date=end_date)
    context["request"] = request
    try:
        pdf_bytes = render_template_to_pdf_bytes(
            templates=templates,
            template_name=context.get("pdf_template_name", "ui/admin/plantilla_duplicado_pdf.html"),
            context=context,
            request=request,
            wkhtmltopdf_args=[
                "--page-size", "Letter",
                "--orientation", "Landscape",
                "--margin-top", "0.35in",
                "--margin-right", "0.45in",
                "--margin-bottom", "0.35in",
                "--margin-left", "0.45in",
            ],
        )
    except PDFBackendUnavailableError as exc:
        return Response(str(exc), status_code=503, media_type="text/plain; charset=utf-8")
    except PDFRenderError as exc:
        return Response(str(exc), status_code=500, media_type="text/plain; charset=utf-8")
    filename = _download_filename("plantilla_duplicado", context, "pdf")
    return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _build_excel_bytes(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Duplicado"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="F4CDAE")
    total_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    ws["B1"] = "COMPAÑIA: Centros Sor Isolina Ferré, Inc."
    ws["B2"] = f"MES REPORTADO: {context['period_label']}"
    program_columns = context.get("program_columns", [])
    headers = ["RESIDENCIAL"] + [program["label"] for program in program_columns] + ["Total Participación", "Participantes No Duplicados", "Total de Servicios"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
    row_index = 5
    for row in context.get("rows", []):
        values = [row["residential_name"]] + [row["programs"].get(program["code"], 0) for program in program_columns] + [row["total_participation"], row["unique_participants"], row["total_services"]]
        for col, value in enumerate(values, start=2):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = left if col == 2 else center
        row_index += 1
    totals = context["duplicado_totals"]
    values = ["Total"] + [totals["programs"].get(program["code"], 0) for program in program_columns] + [totals["total_participation"], totals["unique_participants"], totals["total_services"]]
    for col, value in enumerate(values, start=2):
        cell = ws.cell(row=row_index, column=col, value=value)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
        cell.alignment = center
    for col, width in {"B": 28, "C": 14, "D": 14, "E": 14, "F": 14, "G": 18, "H": 22, "I": 18}.items():
        ws.column_dimensions[col].width = width
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@router.get("/plantilla-duplicado/excel")
def plantilla_duplicado_excel(
    request: Request,
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    residential_id: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    selected_residential_id = _resolve_report_residential_id(
        request,
        current_user,
        _parse_optional_int(residential_id),
    )
    context = _build_context(db, current_user, month, year, _parse_optional_int(proposal_id), selected_residential_id, period_type=period_type, start_date=start_date, end_date=end_date)
    filename = _download_filename("plantilla_duplicado", context, "xlsx")
    return StreamingResponse(BytesIO(_build_excel_bytes(context)), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
