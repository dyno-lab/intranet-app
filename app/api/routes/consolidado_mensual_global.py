from __future__ import annotations

from datetime import date
from io import BytesIO
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, status
from fastapi.responses import RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_db
from app.core.auth import require_admin
from app.models.proposal import Proposal
from app.models.residential import Residential
from app.models.user import User
from app.services.consolidado_mensual_service import (
    MONTH_NAMES,
    build_consolidado_mensual_global,
    build_validation_rows,
)
from app.services.report_pdf import PDFBackendUnavailableError, PDFRenderError, render_template_to_pdf_bytes

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

MONTH_OPTIONS = [(idx, name.capitalize()) for idx, name in MONTH_NAMES.items()]


def _validate_month_year(month: int, year: int) -> None:
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Mes inválido.")
    if year < 2000 or year > 2100:
        raise HTTPException(status_code=400, detail="Año inválido.")


def _validate_period(period_type: str, month: int | None, year: int | None, start_date: str | None, end_date: str | None) -> None:
    if period_type == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="Debe indicar fecha inicial y fecha final.")
        try:
            parsed_start = date.fromisoformat(start_date)
            parsed_end = date.fromisoformat(end_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha inválido.")
        if parsed_start > parsed_end:
            raise HTTPException(status_code=400, detail="La fecha inicial no puede ser mayor que la fecha final.")
        return
    if month is None or year is None:
        raise HTTPException(status_code=400, detail="Debe indicar mes y año.")
    _validate_month_year(month, year)


def _parse_optional_int(value: str | int | None) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Filtro inválido.")


def _base_context(db: Session) -> dict:
    proposals = db.execute(select(Proposal).order_by(Proposal.code)).scalars().all()
    residentials = db.execute(
        select(Residential)
        .where(Residential.is_active == True)  # noqa: E712
        .order_by(Residential.municipality, Residential.name)
    ).scalars().all()
    return {
        "proposals": proposals,
        "residentials": residentials,
        "month_options": MONTH_OPTIONS,
        "period_type_options": [
            {"value": "monthly", "label": "Mensual"},
            {"value": "custom", "label": "Personalizado"},
        ],
        "year_options": list(range(date.today().year - 2, date.today().year + 3)),
    }


def _build_context(
    db: Session,
    current_user: User,
    month: int | None,
    year: int | None,
    proposal_id: int | None,
    residential_id: int | None,
    period_type: str = "monthly",
    start_date: str | None = None,
    end_date: str | None = None,
) -> dict:
    _validate_period(period_type, month, year, start_date, end_date)
    report = build_consolidado_mensual_global(
        db,
        month=month,
        year=year,
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
        proposal_id=proposal_id,
        residential_id=residential_id,
        current_user=current_user,
    )
    return {
        **_base_context(db),
        **report,
        "current_user": current_user,
        "msg": None,
    }


def _query_params(
    month: int | None,
    year: int | None,
    proposal_id: int | None,
    residential_id: int | None,
    period_type: str = "monthly",
    start_date: str | None = None,
    end_date: str | None = None,
) -> str:
    params = {"period_type": period_type}
    if month:
        params["month"] = month
    if year:
        params["year"] = year
    if start_date:
        params["start_date"] = start_date
    if end_date:
        params["end_date"] = end_date
    if proposal_id:
        params["proposal_id"] = proposal_id
    if residential_id:
        params["residential_id"] = residential_id
    return urlencode(params)


def _download_filename(prefix: str, context: dict, extension: str) -> str:
    proposal_code = getattr(context.get("proposal"), "code", None)
    proposal_part = f"_{proposal_code}" if proposal_code else ""
    if context.get("selected_period_type") == "custom":
        suffix = f"{context.get('selected_start_date')}_a_{context.get('selected_end_date')}"
    else:
        suffix = f"{context['month']:02d}_{context['year']}"
    return f"{prefix}{proposal_part}_{suffix}.{extension}"


@router.get("/consolidado-mensual-global")
def consolidado_mensual_global_index(
    request: Request,
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    residential_id: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    selected_month = month or date.today().month
    selected_year = year or date.today().year
    context = _build_context(
        db,
        current_user,
        selected_month,
        selected_year,
        _parse_optional_int(proposal_id),
        _parse_optional_int(residential_id),
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
    )
    context["request"] = request
    return templates.TemplateResponse("ui/admin/consolidado_mensual_global.html", context)


@router.post("/consolidado-mensual-global/generar")
def consolidado_mensual_global_generar(
    month: int | None = Form(None),
    year: int | None = Form(None),
    period_type: str = Form("monthly"),
    start_date: str | None = Form(None),
    end_date: str | None = Form(None),
    proposal_id: str | None = Form(None),
    residential_id: str | None = Form(None),
    output: str = Form("pdf"),
    current_user: User = Depends(require_admin),
):
    _validate_period(period_type, month, year, start_date, end_date)
    proposal_id_int = _parse_optional_int(proposal_id)
    residential_id_int = _parse_optional_int(residential_id)
    target = "pdf" if output == "pdf" else "excel" if output == "excel" else "validacion" if output == "validacion" else ""
    if not target:
        raise HTTPException(status_code=400, detail="Salida inválida.")
    return RedirectResponse(
        f"/ui/admin/consolidado-mensual-global/{target}?{_query_params(month, year, proposal_id_int, residential_id_int, period_type, start_date, end_date)}",
        status_code=status.HTTP_303_SEE_OTHER,
    )


@router.get("/consolidado-mensual-global/pdf")
def consolidado_mensual_global_pdf(
    request: Request,
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    residential_id: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    context = _build_context(
        db,
        current_user,
        month,
        year,
        _parse_optional_int(proposal_id),
        _parse_optional_int(residential_id),
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
    )
    context["request"] = request
    try:
        pdf_bytes = render_template_to_pdf_bytes(
            templates=templates,
            template_name=context.get("pdf_template_name", "ui/admin/consolidado_mensual_global_pdf.html"),
            context=context,
            request=request,
            wkhtmltopdf_args=[
                "--page-size", "Letter",
                "--orientation", "Portrait",
                "--margin-top", "0.35in",
                "--margin-right", "0.45in",
                "--margin-bottom", "0.35in",
                "--margin-left", "0.45in",
            ],
        )
    except PDFBackendUnavailableError as exc:
        return Response(str(exc), status_code=503, media_type="text/plain; charset=utf-8")
    except PDFRenderError as exc:
        return Response(str(exc), status_code=500, media_type="text/plain; charset=utf-8")

    filename = _download_filename("consolidado_mensual_global", context, "pdf")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_excel_bytes(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado Global"
    ws.freeze_panes = "A8"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    total_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:J1")
    ws["A1"] = "Consolidado Mensual Global"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws["A2"] = "Periodo"
    ws["B2"] = context["period_label"]
    ws["A3"] = "Propuesta"
    ws["B3"] = f"{context['proposal'].code} - {context['proposal'].name}" if context.get("proposal") else "Todas"
    ws["A4"] = "Nota"
    ws["B4"] = context["legacy_reference_note"]

    headers = ["Residencial", "Municipio", "RQ", "Programa", "F", "M", "Participantes únicos", "Asistencias/Duplicados"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    row_index = 8
    for row in context.get("rows", []):
        for program in row.get("programs", []):
            values = [
                row["residential_name"],
                row["municipality"],
                row["rq_code"],
                program["label"],
                program["gender"].get("F", 0),
                program["gender"].get("M", 0),
                program["unique_participants"],
                program["attendances"],
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col, value=value)
                cell.border = border
                cell.alignment = left if col <= 4 else center
            row_index += 1

    totals = context["totals"]
    for col, value in enumerate(["TOTAL", "", "", "", totals["gender"]["F"], totals["gender"]["M"], totals["unique_participants"], totals["attendances"]], start=1):
        cell = ws.cell(row=row_index, column=col, value=value)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = border
        cell.alignment = center

    widths = {"A": 28, "B": 18, "C": 12, "D": 28, "E": 10, "F": 10, "G": 20, "H": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@router.get("/consolidado-mensual-global/excel")
def consolidado_mensual_global_excel(
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    residential_id: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    context = _build_context(
        db,
        current_user,
        month,
        year,
        _parse_optional_int(proposal_id),
        _parse_optional_int(residential_id),
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
    )
    filename = _download_filename("consolidado_mensual_global", context, "xlsx")
    return StreamingResponse(
        BytesIO(_build_excel_bytes(context)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/consolidado-mensual-global/validacion")
def consolidado_mensual_global_validacion(
    request: Request,
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    residential_id: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    context = _build_context(
        db,
        current_user,
        month,
        year,
        _parse_optional_int(proposal_id),
        _parse_optional_int(residential_id),
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
    )
    context["request"] = request
    context["validation_rows"] = build_validation_rows(context)
    return templates.TemplateResponse("ui/admin/consolidado_mensual_global_validacion.html", context)
