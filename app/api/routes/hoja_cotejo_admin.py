from __future__ import annotations

from datetime import date
from io import BytesIO
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, status
from fastapi.responses import RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.api.deps import get_db
from app.core.auth import require_admin
from app.models.user import User
from app.services.hoja_cotejo_admin_service import build_hoja_cotejo_admin_context, base_context
from app.services.report_pdf import PDFBackendUnavailableError, PDFRenderError, render_template_to_pdf_bytes

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


def _validate_month_year(month: int, year: int) -> None:
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Mes inválido.")
    if year < 2000 or year > 2100:
        raise HTTPException(status_code=400, detail="Año inválido.")


def _validate_period(period_type: str, month: int | None, year: int | None, start_date: str | None, end_date: str | None) -> None:
    if period_type == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="Debe indicar fecha inicial y fecha final.")
        try:
            parsed_start = date.fromisoformat(start_date)
            parsed_end = date.fromisoformat(end_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha inválido.")
        if parsed_start > parsed_end:
            raise HTTPException(status_code=400, detail="La fecha inicial no puede ser mayor que la fecha final.")
        return
    if month is None or year is None:
        raise HTTPException(status_code=400, detail="Debe indicar mes y año.")
    _validate_month_year(month, year)


def _parse_optional_int(value: str | int | None) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Filtro inválido.")


def _build_context(
    db: Session,
    current_user: User,
    month: int | None,
    year: int | None,
    proposal_id: int | None,
    period_type: str = "monthly",
    start_date: str | None = None,
    end_date: str | None = None,
    authorized_name: str | None = None,
) -> dict:
    _validate_period(period_type, month, year, start_date, end_date)
    return build_hoja_cotejo_admin_context(
        db,
        month=month,
        year=year,
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
        proposal_id=proposal_id,
        authorized_name=authorized_name,
        current_user=current_user,
    )


def _query_params(month: int | None, year: int | None, proposal_id: int | None, period_type: str, start_date: str | None, end_date: str | None, authorized_name: str | None) -> str:
    params = {"period_type": period_type}
    if month:
        params["month"] = month
    if year:
        params["year"] = year
    if proposal_id:
        params["proposal_id"] = proposal_id
    if start_date:
        params["start_date"] = start_date
    if end_date:
        params["end_date"] = end_date
    if authorized_name:
        params["authorized_name"] = authorized_name
    return urlencode(params)


def _download_filename(prefix: str, context: dict, extension: str) -> str:
    proposal_code = getattr(context.get("proposal"), "code", None)
    proposal_part = f"_{proposal_code}" if proposal_code else ""
    if context.get("selected_period_type") == "custom":
        suffix = f"{context.get('selected_start_date')}_a_{context.get('selected_end_date')}"
    else:
        suffix = f"{context['month']:02d}_{context['year']}"
    return f"{prefix}{proposal_part}_{suffix}.{extension}"


@router.get("/hoja-cotejo")
def hoja_cotejo_index(
    request: Request,
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    authorized_name: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    selected_month = month or date.today().month
    selected_year = year or date.today().year
    if not proposal_id:
        context = {**base_context(db), "request": request, "current_user": current_user, "msg": None, "selected_month": selected_month, "selected_year": selected_year, "month": selected_month, "year": selected_year, "selected_period_type": period_type, "selected_start_date": start_date or "", "selected_end_date": end_date or "", "selected_proposal_id": None, "authorized_name": authorized_name or "", "period_label": "", "program_blocks": [], "totals": {"activities_count": 0, "duplicados": 0, "met": 0, "not_met": 0, "rows": 0}}
        return templates.TemplateResponse("ui/admin/hoja_cotejo.html", context)
    context = _build_context(
        db,
        current_user,
        selected_month,
        selected_year,
        _parse_optional_int(proposal_id),
        period_type=period_type,
        start_date=start_date,
        end_date=end_date,
        authorized_name=authorized_name,
    )
    context["request"] = request
    return templates.TemplateResponse("ui/admin/hoja_cotejo.html", context)


@router.post("/hoja-cotejo/generar")
def hoja_cotejo_generar(
    month: int | None = Form(None),
    year: int | None = Form(None),
    period_type: str = Form("monthly"),
    start_date: str | None = Form(None),
    end_date: str | None = Form(None),
    proposal_id: str | None = Form(None),
    authorized_name: str | None = Form(None),
    output: str = Form("pdf"),
    current_user: User = Depends(require_admin),
):
    _validate_period(period_type, month, year, start_date, end_date)
    proposal_id_int = _parse_optional_int(proposal_id)
    if not proposal_id_int:
        raise HTTPException(status_code=400, detail="Debe seleccionar una propuesta.")
    target = "pdf" if output == "pdf" else "excel" if output == "excel" else ""
    if not target:
        raise HTTPException(status_code=400, detail="Salida inválida.")
    return RedirectResponse(
        f"/ui/admin/hoja-cotejo/{target}?{_query_params(month, year, proposal_id_int, period_type, start_date, end_date, authorized_name)}",
        status_code=status.HTTP_303_SEE_OTHER,
    )


@router.get("/hoja-cotejo/pdf")
def hoja_cotejo_pdf(
    request: Request,
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    authorized_name: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    context = _build_context(db, current_user, month, year, _parse_optional_int(proposal_id), period_type=period_type, start_date=start_date, end_date=end_date, authorized_name=authorized_name)
    context["request"] = request
    try:
        pdf_bytes = render_template_to_pdf_bytes(
            templates=templates,
            template_name=context.get("pdf_template_name", "ui/admin/hoja_cotejo_pdf.html"),
            context=context,
            request=request,
            wkhtmltopdf_args=[
                "--page-size", "Letter",
                "--orientation", "Landscape",
                "--margin-top", "0.25in",
                "--margin-right", "0.25in",
                "--margin-bottom", "0.25in",
                "--margin-left", "0.25in",
            ],
        )
    except PDFBackendUnavailableError as exc:
        return Response(str(exc), status_code=503, media_type="text/plain; charset=utf-8")
    except PDFRenderError as exc:
        return Response(str(exc), status_code=500, media_type="text/plain; charset=utf-8")
    filename = _download_filename("hoja_cotejo", context, "pdf")
    return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _build_excel_bytes(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja Cotejo"
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    program_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A1"] = "HOJA MENSUAL DE COTEJO DE PROGRAMAS LOGRADAS POR ACTIVIDAD SEGÚN EL PLAN DE TRABAJO"
    ws["A2"] = f"PERÍODO: {context['period_label']}"
    ws["A3"] = f"PROPUESTA: {context['proposal'].code} - {context['proposal'].name}" if context.get("proposal") else "PROPUESTA:"
    row_index = 5
    headers = ["Programa", "Actividad", "Descripción", "Realizadas", "Duplicados / Personas impactadas", "Sí", "No", "Cumplimiento", "% Cumplimiento mensual", "Actividades logradas por periodo propuesta", "% Cumplimiento acumulado"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_index, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
    row_index += 1
    for program in context.get("program_blocks", []):
        for item in program.get("rows", []):
            values = [program["program_display_name"], item["activity_code"], item["activity_description"], item["activities_count"], item["duplicados"], "X" if item["met"] else "", "" if item["met"] else "X", item["goal_summary"], f"{item['monthly_percent']}%" if item["monthly_percent"] is not None else "N/A", item["cumulative_ratio"], f"{item['percent']}%" if item["percent"] is not None else "N/A"]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col, value=value)
                cell.border = border
                cell.alignment = left if col <= 3 or col == 8 else center
            row_index += 1
    for col, width in {"A": 22, "B": 16, "C": 54, "D": 14, "E": 22, "F": 8, "G": 8, "H": 28, "I": 20, "J": 24, "K": 20}.items():
        ws.column_dimensions[col].width = width
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@router.get("/hoja-cotejo/excel")
def hoja_cotejo_excel(
    month: int | None = Query(None),
    year: int | None = Query(None),
    period_type: str = Query("monthly"),
    start_date: str | None = Query(None),
    end_date: str | None = Query(None),
    proposal_id: str | None = Query(None),
    authorized_name: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    context = _build_context(db, current_user, month, year, _parse_optional_int(proposal_id), period_type=period_type, start_date=start_date, end_date=end_date, authorized_name=authorized_name)
    filename = _download_filename("hoja_cotejo", context, "xlsx")
    return StreamingResponse(BytesIO(_build_excel_bytes(context)), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
