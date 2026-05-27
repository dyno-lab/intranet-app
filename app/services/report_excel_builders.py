from __future__ import annotations

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
META_LABEL_FILL = PatternFill(fill_type="solid", fgColor="F3F6FA")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
SUBTITLE_FONT = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_title(ws, cell_ref: str, value: str, merge_to: str | None = None):
    ws[cell_ref] = value
    cell = ws[cell_ref]
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if merge_to:
        ws.merge_cells(f"{cell_ref}:{merge_to}")
        for row in ws[cell_ref:merge_to]:
            for merged_cell in row:
                merged_cell.fill = TITLE_FILL
                merged_cell.border = THIN_BORDER


def style_subtitle(ws, cell_ref: str, value: str, merge_to: str | None = None):
    ws[cell_ref] = value
    cell = ws[cell_ref]
    cell.font = SUBTITLE_FONT
    cell.alignment = LEFT
    if merge_to:
        ws.merge_cells(f"{cell_ref}:{merge_to}")


def style_meta_label(cell):
    cell.font = HEADER_FONT
    cell.fill = META_LABEL_FILL
    cell.border = THIN_BORDER
    cell.alignment = LEFT


def style_meta_value(cell):
    cell.border = THIN_BORDER
    cell.alignment = LEFT


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = CENTER


def style_section(cell):
    cell.font = Font(bold=True, size=12)
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER
    cell.alignment = LEFT


def style_total(cell):
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER
    cell.alignment = CENTER


def excel_percent(value: float | int | None) -> float:
    if value is None:
        return 0
    return float(value) / 100


def apply_table_border(ws, start_row: int, end_row: int, start_col: int, end_col: int, left_align_first_col: bool = True):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col == start_col and left_align_first_col:
                cell.alignment = LEFT
            elif cell.alignment == Alignment():
                cell.alignment = CENTER


def make_workbook() -> Workbook:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    return wb


def make_sheet(wb: Workbook, title: str):
    ws = wb.create_sheet(title=title[:31])
    ws.sheet_view.showGridLines = False
    return ws


def proposal_label_from_context(context: dict) -> str:
    return next(
        (f"{proposal.code} - {proposal.name}" for proposal in context.get("proposals", []) if proposal.proposal_id == context.get("selected_proposal_id")),
        "",
    )


def build_bonafide_sheet(wb: Workbook, context: dict, title: str = "Bonafide"):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "A8"
    style_title(ws, "A1", "Programa Faro de Esperanza", "H1")
    style_subtitle(ws, "A2", "Listado Bonafide", "H2")

    meta = [("Periodo", context["period_label"]), ("Residencial", context["residential_name"] or ""), ("Municipio", context["municipality"] or "")]
    for idx, (label, value) in enumerate(meta, start=4):
        style_meta_label(ws.cell(row=idx, column=1, value=label))
        style_meta_value(ws.cell(row=idx, column=2, value=value))

    for col_index, header in enumerate(["#", "Expediente", "Nombre", "F", "M", "Edad", "Edif.", "Apto."], start=1):
        style_header(ws.cell(row=8, column=col_index, value=header))

    end_row = 8
    for row_index, row in enumerate(context.get("rows", []), start=9):
        ws.cell(row=row_index, column=1, value=row.get("index", "")).alignment = CENTER
        ws.cell(row=row_index, column=2, value=row.get("expediente", "")).alignment = LEFT
        ws.cell(row=row_index, column=3, value=row.get("nombre", "")).alignment = LEFT
        ws.cell(row=row_index, column=4, value=row.get("f", "")).alignment = CENTER
        ws.cell(row=row_index, column=5, value=row.get("m", "")).alignment = CENTER
        ws.cell(row=row_index, column=6, value=row.get("edad", "")).alignment = CENTER
        ws.cell(row=row_index, column=7, value=row.get("edificio", "")).alignment = CENTER
        ws.cell(row=row_index, column=8, value=row.get("apartamento", "")).alignment = CENTER
        end_row = row_index

    if end_row == 8:
        ws.cell(row=9, column=1, value="No hay datos para este filtro.")
        ws.merge_cells("A9:H9")
        ws["A9"].alignment = CENTER
        end_row = 9

    apply_table_border(ws, 8, end_row, 1, 8)
    for col, width in {"A": 6, "B": 20, "C": 40, "D": 6, "E": 6, "F": 8, "G": 12, "H": 12}.items():
        ws.column_dimensions[col].width = width
    return ws


def build_no_duplicado_sheet(wb: Workbook, context: dict, title: str = "No Duplicado", duplicated: bool = False):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "A10"
    style_title(ws, "A1", "Informe mensual de participaciones" if duplicated else "Informe mensual de participantes", "D1")
    style_subtitle(ws, "A2", "Duplicado por edad y sexo en los proyectos impactados" if duplicated else "No Duplicado por edad y sexo en los proyectos impactados", "D2")

    meta = [
        ("Residencial", context["residential_name"] or ""),
        ("Municipio", context["municipality"] or ""),
        ("RQ", context["rq_code"] or ""),
        ("Periodo reportado", context["period_label"]),
        ("Funcionario autorizado", context["authorized_name"] or ""),
    ]
    for idx, (label, value) in enumerate(meta, start=4):
        style_meta_label(ws.cell(row=idx, column=1, value=label))
        style_meta_value(ws.cell(row=idx, column=2, value=value))

    headers = ["Clasificación", "F", "M", "Total de participaciones" if duplicated else "Total de participantes"]
    for col_index, header in enumerate(headers, start=1):
        style_header(ws.cell(row=10, column=col_index, value=header))

    row_index = 11
    for row in context.get("rows", []):
        ws.cell(row=row_index, column=1, value=row.get("label", "")).alignment = LEFT
        ws.cell(row=row_index, column=2, value=row.get("f", 0)).alignment = CENTER
        ws.cell(row=row_index, column=3, value=row.get("m", 0)).alignment = CENTER
        ws.cell(row=row_index, column=4, value=row.get("total", 0)).alignment = CENTER
        row_index += 1

    for col_index, value in enumerate(["TOTAL", context["total_f"], context["total_m"], context["total_all"]], start=1):
        cell = ws.cell(row=row_index, column=col_index, value=value)
        style_total(cell)
        if col_index == 1:
            cell.alignment = LEFT

    apply_table_border(ws, 10, row_index, 1, 4)
    for col, width in {"A": 35, "B": 10, "C": 10, "D": 20}.items():
        ws.column_dimensions[col].width = width
    return ws


def build_visitas_sheet(wb: Workbook, context: dict, title: str = "Visitas", rows: list[dict] | None = None, include_totals_when_empty: bool = False):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "A6"
    style_title(ws, "A1", "CENTROS SOR ISOLINA FERRÉ", "K1")
    style_subtitle(ws, "A2", "Reporte de Visitas", "K2")
    meta = [
        ("Propuesta", proposal_label_from_context(context)),
        ("Periodo", context["period_label"]),
        ("Residencial", context["residential_name"] or ""),
        ("Visitas", context["summary"]["visits"]),
        ("Asistencias", context["summary"]["attendances"]),
        ("Horas", context["summary"]["hours"]),
    ]
    positions = [(3,1),(3,4),(4,1),(4,4),(4,7),(4,10)]
    for (row, col), (label, value) in zip(positions, meta):
        style_meta_label(ws.cell(row=row, column=col, value=label))
        style_meta_value(ws.cell(row=row, column=col+1, value=value))
        if label == "Horas":
            ws.cell(row=row, column=col+1).number_format = "0.00"

    for col_index, header in enumerate(["Empleado", "Visitas registradas", "Asistencias acumuladas", "Horas acumuladas"], start=1):
        style_header(ws.cell(row=6, column=col_index, value=header))

    sheet_rows = rows if rows is not None else context.get("rows", [])
    row_index = 7
    for row in sheet_rows:
        values = [row.get("employee_name", ""), row.get("visits", 0), row.get("attendances", 0), row.get("hours", 0)]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if col_index == 4:
                cell.number_format = "0.00"
            cell.alignment = LEFT if col_index == 1 else CENTER
        row_index += 1

    if sheet_rows or include_totals_when_empty:
        totals = ["TOTALES", context["summary"]["visits"], context["summary"]["attendances"], context["summary"]["hours"]]
        for col_index, value in enumerate(totals, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            style_total(cell)
            if col_index == 1:
                cell.alignment = LEFT
            if col_index == 4:
                cell.number_format = "0.00"

    apply_table_border(ws, 6, row_index, 1, 4)
    ws.auto_filter.ref = f"A6:D{max(row_index, 6)}"
    for col, width in {"A": 34, "B": 18, "C": 20, "D": 18}.items():
        ws.column_dimensions[col].width = width
    return ws


def build_por_programa_sheet(wb: Workbook, context: dict, title: str = "Por Programa"):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "A10"
    style_title(ws, "A1", "Informe mensual de participantes", "D1")
    style_subtitle(ws, "A2", "Participación no duplicada por programa, edad y sexo en los proyectos impactados", "D2")

    meta = [
        ("Residencial", context["residential_name"] or ""),
        ("Municipio", context["municipality"] or ""),
        ("RQ", context["rq_code"] or ""),
        ("Periodo reportado", context["period_label"]),
        ("Funcionario autorizado", context["authorized_name"] or ""),
    ]
    for idx, (label, value) in enumerate(meta, start=4):
        style_meta_label(ws.cell(row=idx, column=1, value=label))
        style_meta_value(ws.cell(row=idx, column=2, value=value))

    row_index = 10
    for section in context.get("program_sections", []):
        style_section(ws.cell(row=row_index, column=1, value=f"Programa: {section['program_display_name']}"))
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)

        ws.cell(row=row_index + 1, column=1, value="Actividades adjudicadas")
        ws.cell(row=row_index + 1, column=2, value=section["assigned_activity_count"])
        style_meta_label(ws.cell(row=row_index + 1, column=1, value="Actividades adjudicadas"))
        style_meta_value(ws.cell(row=row_index + 1, column=2, value=section["assigned_activity_count"]))

        header_row = row_index + 3
        style_header(ws.cell(row=header_row, column=1, value="Clasificación"))
        style_header(ws.cell(row=header_row, column=2, value="Número de participantes por edad y sexo"))
        style_header(ws.cell(row=header_row, column=4, value="Total de participantes"))
        ws.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=3)
        ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row + 1, end_column=1)
        ws.merge_cells(start_row=header_row, start_column=4, end_row=header_row + 1, end_column=4)
        style_header(ws.cell(row=header_row + 1, column=2, value="F"))
        style_header(ws.cell(row=header_row + 1, column=3, value="M"))
        ws.cell(row=header_row + 1, column=1).border = THIN_BORDER
        ws.cell(row=header_row + 1, column=4).border = THIN_BORDER

        current_row = header_row + 2
        for row in section.get("rows", []):
            ws.cell(row=current_row, column=1, value=row["label"]).alignment = LEFT
            ws.cell(row=current_row, column=2, value=row["f"]).alignment = CENTER
            ws.cell(row=current_row, column=3, value=row["m"]).alignment = CENTER
            ws.cell(row=current_row, column=4, value=row["total"]).alignment = CENTER
            current_row += 1

        for col_index, value in enumerate(["TOTAL", section["total_f"], section["total_m"], section["total_all"]], start=1):
            cell = ws.cell(row=current_row, column=col_index, value=value)
            style_total(cell)
            if col_index == 1:
                cell.alignment = LEFT

        apply_table_border(ws, header_row, current_row, 1, 4)
        row_index = current_row + 3

    for col, width in {"A": 40, "B": 14, "C": 14, "D": 22}.items():
        ws.column_dimensions[col].width = width
    return ws


def _hoja_cotejo_cell_value(row: dict, population_label: str, column: dict):
    key = column.get("key")
    if key == "population_label":
        return population_label
    if key == "activity_text":
        return f"{row.get('activity_code', '')} {row.get('activity_description') or ''}".strip()
    return row.get(key, "")


def build_hoja_cotejo_sheet(wb: Workbook, context: dict, title: str = "Hoja de Cotejo"):
    ws = make_sheet(wb, title)
    columns = context.get("report_template_columns") or [
        {"key": "activity_text", "label": "Actividad", "align": "left"},
        {"key": "activities_count", "label": "Realizadas", "align": "center"},
        {"key": "duplicados", "label": "Duplicados", "align": "center"},
        {"key": "unique_participants", "label": "Unicos", "align": "center"},
        {"key": "contact_hours", "label": "Horas contacto", "align": "center", "format": "decimal_2"},
    ]
    end_col = max(len(columns), 5)
    ws.freeze_panes = "A10"
    style_title(ws, "A1", "Hoja de Cotejo", ws.cell(row=1, column=end_col).coordinate)
    style_subtitle(ws, "A2", "Reporte por programa, clasificacion y actividad", ws.cell(row=2, column=end_col).coordinate)
    meta = [
        ("Residencial", context["residential_name"] or ""),
        ("Municipio", context["municipality"] or ""),
        ("RQ", context["rq_code"] or ""),
        ("Periodo reportado", context["period_label"]),
    ]
    for idx, (label, value) in enumerate(meta, start=4):
        style_meta_label(ws.cell(row=idx, column=1, value=label))
        style_meta_value(ws.cell(row=idx, column=2, value=value))

    row_index = 10
    for program_block in context.get("program_blocks", []):
        style_section(ws.cell(row=row_index, column=1, value=f"Programa: {program_block['program_display_name']}"))
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(columns))
        row_index += 1
        for population_block in program_block.get("population_blocks", []):
            style_meta_label(ws.cell(row=row_index, column=1, value=f"Clasificacion / poblacion: {population_block['population_label']}"))
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(columns))
            row_index += 1
            for col_index, column in enumerate(columns, start=1):
                style_header(ws.cell(row=row_index, column=col_index, value=column.get("label", "")))
            header_row = row_index
            row_index += 1

            if population_block.get("rows"):
                for row in population_block["rows"]:
                    for col_index, column in enumerate(columns, start=1):
                        cell = ws.cell(
                            row=row_index,
                            column=col_index,
                            value=_hoja_cotejo_cell_value(row, population_block["population_label"], column),
                        )
                        cell.alignment = LEFT if column.get("align") == "left" or col_index == 1 else CENTER
                        if column.get("format") == "decimal_2":
                            cell.number_format = "0.00"
                    row_index += 1
            else:
                ws.cell(row=row_index, column=1, value="No hay actividades asignadas a esta clasificacion.")
                ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(columns))
                ws.cell(row=row_index, column=1).alignment = CENTER
                row_index += 1
            apply_table_border(ws, header_row, row_index - 1, 1, len(columns))
            row_index += 1
        row_index += 1

    style_total(ws.cell(row=row_index, column=1, value="Total Horas Contacto"))
    total_cell = ws.cell(row=row_index, column=2, value=context["total_contact_hours"])
    style_total(total_cell)
    total_cell.number_format = "0.00"
    for col_index, column in enumerate(columns, start=1):
        letter = get_column_letter(col_index)
        width_hint = str(column.get("width", "")).strip().replace("%", "")
        try:
            ws.column_dimensions[letter].width = max(12, min(60, float(width_hint) * 1.4))
        except ValueError:
            ws.column_dimensions[letter].width = 16 if col_index > 1 else 32
    return ws

def build_desercion_sheet(wb: Workbook, context: dict, title: str = "Desercion"):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "B6"
    style_title(ws, "A1", "CENTROS SOR ISOLINA FERRÉ", "Z1")
    style_subtitle(ws, "A2", "Informe de Deserción Escolar", "Z2")
    meta = [
        ("Propuesta", proposal_label_from_context(context)),
        ("Periodo", context["period_label"]),
        ("Residencial", context["residential_name"] or ""),
        ("Reclutados totales", context["total"]["recruited"]),
    ]
    positions = [(3,1),(3,4),(4,1),(4,4)]
    for (row, col), (label, value) in zip(positions, meta):
        style_meta_label(ws.cell(row=row, column=col, value=label))
        style_meta_value(ws.cell(row=row, column=col+1, value=value))

    headers = ["Residencial", "Total", "F", "M", *context["grade_columns"], "Tutorías", "% Tutorías", "Escuela", "% Escuela", "10", "20", "30", "40"]
    pct_cols = {len(context["grade_columns"]) + 6, len(context["grade_columns"]) + 8}
    for col_index, header in enumerate(headers, start=1):
        style_header(ws.cell(row=6, column=col_index, value=header))

    row_index = 7
    for row in context.get("rows", []):
        values = [
            row["residential_name"], row["recruited"], row["f"], row["m"],
            *[row["grades"].get(grade, 0) for grade in context["grade_columns"]],
            row["tutoring"], row["tutoring_pct"] / 100, row["school"], row["school_pct"] / 100,
            row["report_10"], row["report_20"], row["report_30"], row["report_40"],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = LEFT if col_index == 1 else CENTER
            if col_index in pct_cols:
                cell.number_format = "0.00%"
        row_index += 1

    total_values = [
        "TOTAL", context["total"]["recruited"], context["total"]["f"], context["total"]["m"],
        *[context["total"]["grades"].get(grade, 0) for grade in context["grade_columns"]],
        context["total"]["tutoring"], context["total"]["tutoring_pct"] / 100,
        context["total"]["school"], context["total"]["school_pct"] / 100,
        context["total"]["report_10"], context["total"]["report_20"], context["total"]["report_30"], context["total"]["report_40"],
    ]
    for col_index, value in enumerate(total_values, start=1):
        cell = ws.cell(row=row_index, column=col_index, value=value)
        style_total(cell)
        if col_index == 1:
            cell.alignment = LEFT
        if col_index in pct_cols:
            cell.number_format = "0.00%"

    apply_table_border(ws, 6, row_index, 1, len(headers))
    widths = {
        "A": 26, "B": 10, "C": 8, "D": 8, "E": 6, "F": 6, "G": 6, "H": 6, "I": 6,
        "J": 6, "K": 6, "L": 6, "M": 6, "N": 6, "O": 6, "P": 6, "Q": 6, "R": 6,
        "S": 10, "T": 12, "U": 10, "V": 12, "W": 8, "X": 8, "Y": 8, "Z": 8,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    return ws


def build_embarazo_sheet(wb: Workbook, context: dict, title: str = "Embarazo"):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "A6"
    style_title(ws, "A1", "CENTROS SOR ISOLINA FERRÉ", "I1")
    style_subtitle(ws, "A2", "Informe de Embarazo", "I2")
    meta = [
        ("Propuesta", proposal_label_from_context(context)),
        ("Periodo", context["period_label"]),
        ("Residencial", context["residential_name"] or ""),
        ("Participación total", context["total"]["participation"]),
    ]
    positions = [(3,1),(3,4),(4,1),(4,4)]
    for (row, col), (label, value) in zip(positions, meta):
        style_meta_label(ws.cell(row=row, column=col, value=label))
        style_meta_value(ws.cell(row=row, column=col+1, value=value))

    headers = [
        "Residencial", "Total reclutados", "F", "M", "Participantes femeninas embarazadas",
        "Participantes masculinos que han embarazado", "% Prevención", "Embarazos", "No embarazos",
    ]
    for col_index, header in enumerate(headers, start=1):
        style_header(ws.cell(row=6, column=col_index, value=header))

    row_index = 7
    for row in context.get("rows", []):
        values = [
            row["residential_name"], row["recruited"], row["f"], row["m"], row["pregnant_f"],
            row["pregnant_m"], row["prevention_pct"] / 100, row["pregnancy_cases"], row["non_pregnant"],
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = LEFT if col_index == 1 else CENTER
            if col_index == 7:
                cell.number_format = "0.00%"
        row_index += 1

    total_values = [
        "TOTAL", context["total"]["recruited"], context["total"]["f"], context["total"]["m"],
        context["total"]["pregnant_f"], context["total"]["pregnant_m"], context["total"]["prevention_pct"] / 100,
        context["total"]["pregnancy_cases"], context["total"]["non_pregnant"],
    ]
    for col_index, value in enumerate(total_values, start=1):
        cell = ws.cell(row=row_index, column=col_index, value=value)
        style_total(cell)
        if col_index == 1:
            cell.alignment = LEFT
        if col_index == 7:
            cell.number_format = "0.00%"

    apply_table_border(ws, 6, row_index, 1, 9)
    ws.auto_filter.ref = f"A6:I{max(row_index, 6)}"
    for col, width in {"A": 28, "B": 15, "C": 8, "D": 8, "E": 22, "F": 24, "G": 14, "H": 12, "I": 14}.items():
        ws.column_dimensions[col].width = width
    return ws


def build_notas_sheet(wb: Workbook, context: dict, title: str = "Notas", include_residential: bool = True, include_subjects: bool = True):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "A6"
    style_title(ws, "A1", "CENTROS SOR ISOLINA FERRÉ", "I1")
    style_subtitle(ws, "A2", "Informe de Notas", "I2")
    meta = [
        ("Propuesta", context.get("proposal_label", proposal_label_from_context(context))),
        ("Periodo", context["period_label"]),
        ("Residencial", context["residential_name"] or ""),
        ("Total evaluados", context["total_row"]["TOTAL"]),
    ]
    positions = [(3,1),(3,4),(4,1),(4,4)]
    for (row, col), (label, value) in zip(positions, meta):
        style_meta_label(ws.cell(row=row, column=col, value=label))
        style_meta_value(ws.cell(row=row, column=col+1, value=value))

    for col_index, header in enumerate(["Nota", "Cantidad", "%"], start=1):
        style_header(ws.cell(row=6, column=col_index, value=header))

    row_index = 7
    summary_start = row_index
    for segment in context.get("general_chart_segments", []):
        ws.cell(row=row_index, column=1, value=segment["label"]).alignment = LEFT
        ws.cell(row=row_index, column=2, value=segment["value"]).alignment = CENTER
        pct_cell = ws.cell(row=row_index, column=3, value=segment["percentage"] / 100)
        pct_cell.number_format = "0.00%"
        pct_cell.alignment = CENTER
        row_index += 1
    apply_table_border(ws, 6, max(row_index - 1, 6), 1, 3)

    row_index += 1
    detail_header = row_index
    for col_index, header in enumerate(["Edad", "A", "B", "C", "D", "F", "Especial", "K", "TOTAL"], start=1):
        style_header(ws.cell(row=row_index, column=col_index, value=header))
    row_index += 1

    for row in context.get("rows", []):
        values = [row["age_label"], row["A"], row["B"], row["C"], row["D"], row["F"], row["Especial"], row["K"], row["TOTAL"]]
        for col_index, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col_index, value=value).alignment = LEFT if col_index == 1 else CENTER
        row_index += 1

    total_values = ["TOTALES", context["total_row"]["A"], context["total_row"]["B"], context["total_row"]["C"], context["total_row"]["D"], context["total_row"]["F"], context["total_row"]["Especial"], context["total_row"]["K"], context["total_row"]["TOTAL"]]
    for col_index, value in enumerate(total_values, start=1):
        cell = ws.cell(row=row_index, column=col_index, value=value)
        style_total(cell)
        if col_index == 1:
            cell.alignment = LEFT
    apply_table_border(ws, detail_header, row_index, 1, 9)

    if include_residential:
        row_index += 2
        residential_header = row_index
        style_section(ws.cell(row=row_index, column=1, value="Distribución por residencial"))
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        row_index += 1
        for col_index, header in enumerate(["Residencial", "A", "B", "C", "D", "F", "TOTAL"], start=1):
            style_header(ws.cell(row=row_index, column=col_index, value=header))
        row_index += 1

        for residential_row in context.get("residential_chart_rows", []):
            values = [
                residential_row["residential_name"], residential_row["A"], residential_row["B"], residential_row["C"],
                residential_row["D"], residential_row["F"], residential_row["total"],
            ]
            for col_index, value in enumerate(values, start=1):
                ws.cell(row=row_index, column=col_index, value=value).alignment = LEFT if col_index == 1 else CENTER
            row_index += 1
        apply_table_border(ws, residential_header + 1, max(row_index - 1, residential_header + 1), 1, 7)

    if include_subjects:
        row_index += 2
        subject_header = row_index
        style_section(ws.cell(row=row_index, column=1, value="Distribución por materia"))
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=6)
        row_index += 1
        for col_index, header in enumerate(["Materia", "A", "B", "C", "D", "F"], start=1):
            style_header(ws.cell(row=row_index, column=col_index, value=header))
        row_index += 1

        for subject_card in context.get("subject_chart_cards", []):
            values = [
                subject_card["subject_name"], subject_card["counts"]["A"], subject_card["counts"]["B"],
                subject_card["counts"]["C"], subject_card["counts"]["D"], subject_card["counts"]["F"],
            ]
            for col_index, value in enumerate(values, start=1):
                ws.cell(row=row_index, column=col_index, value=value).alignment = LEFT if col_index == 1 else CENTER
            row_index += 1
        apply_table_border(ws, subject_header + 1, max(row_index - 1, subject_header + 1), 1, 6)

    for col, width in {"A": 24, "B": 16, "C": 12, "D": 16, "E": 16, "F": 12, "G": 12, "H": 12, "I": 12}.items():
        ws.column_dimensions[col].width = width
    return ws


def build_vca_sheet(wb: Workbook, context: dict, title: str = "VCA"):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "A8"
    style_title(ws, "A1", "ÁREA DE PROGRAMAS COMUNALES Y DE RESIDENTES", "H1")
    style_subtitle(ws, "A2", "INFORME VCA", "H2")
    meta = [
        ("Propuesta", proposal_label_from_context(context)),
        ("Residencial", context["residential_name"] or ""),
        ("Periodo reportado", context["period_label"]),
        ("Total personas con impedimentos", context["total_people"]),
    ]
    for idx, (label, value) in enumerate(meta, start=3):
        style_meta_label(ws.cell(row=idx, column=1, value=label))
        style_meta_value(ws.cell(row=idx, column=2, value=value))

    headers = ["Expediente", "Nombre", "Género", "Edad", *[column.name for column in context.get("columns", [])]]
    for col_index, header in enumerate(headers, start=1):
        style_header(ws.cell(row=8, column=col_index, value=header))

    end_row = 8
    for row_index, row in enumerate(context.get("rows", []), start=9):
        ws.cell(row=row_index, column=1, value=row.get("expediente", "")).alignment = LEFT
        ws.cell(row=row_index, column=2, value=row.get("nombre", "")).alignment = LEFT
        ws.cell(row=row_index, column=3, value=row.get("genero", "")).alignment = CENTER
        ws.cell(row=row_index, column=4, value=row.get("edad", "")).alignment = CENTER
        for offset, column in enumerate(context.get("columns", []), start=5):
            ws.cell(row=row_index, column=offset, value=row["column_values"].get(column.vca_column_id, "")).alignment = LEFT
        end_row = row_index

    apply_table_border(ws, 8, end_row, 1, max(4, len(headers)))
    ws.auto_filter.ref = f"A8:{chr(64 + len(headers))}{max(end_row, 8)}" if len(headers) <= 26 else None
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    for index in range(len(context.get("columns", []))):
        ws.column_dimensions[chr(69 + index)].width = 28
    return ws


def build_adm_sheet(wb: Workbook, context: dict, title: str = "ADM"):
    ws = make_sheet(wb, title)
    ws.freeze_panes = "A7"
    style_title(ws, "A1", "AREA DE PROGRAMAS COMUNALES Y DE RESIDENTES", "E1")
    style_subtitle(ws, "A2", "INFORME ADM", "E2")
    meta = [
        ("Propuesta", proposal_label_from_context(context)),
        ("Residencial", context["residential_name"] or ""),
        ("Periodo reportado", context["period_label"]),
        ("Funcionario autorizado", context["authorized_name"] or ""),
    ]
    for idx, (label, value) in enumerate(meta, start=3):
        style_meta_label(ws.cell(row=idx, column=1, value=label))
        style_meta_value(ws.cell(row=idx, column=2, value=value))

    for idx, header in enumerate(["Tipo de Servicio", "Servicios", "Duplicados", "No Duplicados"], start=1):
        style_header(ws.cell(row=7, column=idx, value=header))

    row_index = 8
    for row in context.get("rows", []):
        ws.cell(row=row_index, column=1, value=row["service_type_name"]).alignment = LEFT
        ws.cell(row=row_index, column=2, value=row["services_count"]).alignment = CENTER
        ws.cell(row=row_index, column=3, value=row["duplicados"]).alignment = CENTER
        ws.cell(row=row_index, column=4, value=row["no_duplicados"]).alignment = CENTER
        row_index += 1

    if not context.get("rows"):
        ws.cell(row=row_index, column=1, value="No hay tipos de servicio ADM configurados o no hay datos para ese filtro.")
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
        ws.cell(row=row_index, column=1).alignment = CENTER
        row_index += 1
    apply_table_border(ws, 7, row_index - 1, 1, 4)

    row_index += 2
    style_section(ws.cell(row=row_index, column=1, value="Socio-Demográfico"))
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
    row_index += 1
    for idx, header in enumerate(["Edad", "F", "M", "Por Ciento", "Diversidad Funcional"], start=1):
        style_header(ws.cell(row=row_index, column=idx, value=header))
    section_start = row_index
    row_index += 1
    for row in context.get("sociodemographic_rows", []):
        ws.cell(row=row_index, column=1, value=row["label"]).alignment = LEFT
        ws.cell(row=row_index, column=2, value=row["f"]).alignment = CENTER
        ws.cell(row=row_index, column=3, value=row["m"]).alignment = CENTER
        pct = ws.cell(row=row_index, column=4, value=excel_percent(row.get("percent")))
        pct.number_format = "0.00%"
        pct.alignment = CENTER
        ws.cell(row=row_index, column=5, value=row["vca"]).alignment = CENTER
        row_index += 1
    total_pct = ws.cell(row=row_index, column=4, value=1 if context["sociodemographic_total"]["total"] else 0)
    for col_index, value in enumerate(["TOTAL", context["sociodemographic_total"]["f"], context["sociodemographic_total"]["m"], 1 if context["sociodemographic_total"]["total"] else 0, context["sociodemographic_total"]["vca"]], start=1):
        cell = ws.cell(row=row_index, column=col_index, value=value)
        style_total(cell)
        if col_index == 1:
            cell.alignment = LEFT
        if col_index == 4:
            cell.number_format = "0.00%"
    apply_table_border(ws, section_start, row_index, 1, 5)

    row_index += 2
    style_section(ws.cell(row=row_index, column=1, value="Composición Familiar"))
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
    row_index += 1
    for idx, header in enumerate(["Composición familiar", "Cantidad"], start=1):
        style_header(ws.cell(row=row_index, column=idx, value=header))
    family_start = row_index
    row_index += 1
    for row in context.get("family_rows", []):
        ws.cell(row=row_index, column=1, value=row["label"]).alignment = LEFT
        ws.cell(row=row_index, column=2, value=row["count"]).alignment = CENTER
        row_index += 1
    for col_index, value in enumerate(["TOTAL", context["family_total"]], start=1):
        cell = ws.cell(row=row_index, column=col_index, value=value)
        style_total(cell)
        if col_index == 1:
            cell.alignment = LEFT
    apply_table_border(ws, family_start, row_index, 1, 2)

    for col, width in {"A": 40, "B": 14, "C": 14, "D": 16, "E": 22}.items():
        ws.column_dimensions[col].width = width
    return ws


def workbook_to_bytes(wb: Workbook) -> BytesIO:
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
